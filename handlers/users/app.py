import msoffcrypto
import openpyxl
from io import BytesIO

async def decrypt_excel(file_path):
    password="1112"
    decrypted = BytesIO()
    with open(file_path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    return decrypted

def get_formula_value(sheet, cell_address):
    try:
        value = sheet[cell_address].value
        return value if value is not None else "N/A"
    except Exception as e:
        print(f"Error reading cell {cell_address}: {e}")
        return None

async def search_in_sheet_by_passport_pin(course, passport_pin, file_path):
    print(f"Searching in course: {course} for passport pin: {passport_pin}")

    try:
        decrypted_file = await decrypt_excel(file_path)
    except Exception as e:
        print(f"Failed to decrypt Excel file: {e}")
        return None

    workbook = openpyxl.load_workbook(decrypted_file, data_only=True)  # Load workbook with formulas calculated

    if course in workbook.sheetnames:
        sheet = workbook[course]
        print(sheet)
    else:
        print(f"Sheet '{course}' not found in the workbook.")
        return None

    found = False
    print(f"Searching for '{passport_pin}' in column C...")  # Adjust column if needed

    # Search through the rows in the selected sheet
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row[2] is not None:  # Check if the cell is not None
            if str(row[2]) == str(passport_pin):  # Compare passport_pin in column C
                found = True
                print(f"Found '{passport_pin}' in column C.")
                cell=row[25]
                print("his ", str(cell))
                data = {
                    "F.I.Sh": row[1],  # Second column (B)
                    "JSHSHIR": row[2],  # Third column (C)
                    "Kurs": row[3],  # Fourth column (D)
                    "Yo'nalishi": row[4],  # Fifth column (E)
                    "Guruh": row[5],  # Sixth column (F)
                    "Ta'lim turi": row[6],  # Eighth column (H)
                    "Ta'lim shakli": row[7],  # Seventh column (G),
                    "kontrakt": str(row[24]),  # Evaluate formula in J93
                    "tolov": str(row[25]),  # Evaluate formula in L93
                }
                print("Data found at row index:", index)
                print("Data extracted:", data)
                return data

    if not found:
        print(f"'{passport_pin}' not found in column C.")
    return None


async def search_in_sheet_by_passport_pin_payment(course, passport_pin, file_path):
    try:
        decrypted_file = await decrypt_excel(file_path)
    except Exception as e:
        print(f"Failed to decrypt Excel file: {e}")
        return None

    workbook = openpyxl.load_workbook(decrypted_file, data_only=True)  # Load workbook with formulas calculated

    if course in workbook.sheetnames:
        sheet = workbook[course]
        print(sheet)
    else:
        print(f"Sheet '{course}' not found in the workbook.")
        return None

    found = False
    print(f"Searching for '{passport_pin}' in column C...")  # Adjust column if needed

    # Search through the rows in the selected sheet
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if str(row[2]) == str(passport_pin):  # Compare passport_pin in column C
            found = True
            print(f"Found '{passport_pin}' in column C.")
            cell = row[25]
            print("his ", str(cell))
            data = {
                "F.I.Sh": row[1],  # Second column (B)
                "JSHSHIR": row[2],  # Third column (C)
                "Kurs": row[3],  # Fourth column (D)
                "Yo'nalishi": row[4],  # Fifth column (E)
                "Guruh": row[5],  # Sixth column (F)
                "Ta'lim turi": row[6],  # Eighth column (H)
                "Ta'lim shakli": row[7],  # Seventh column (G),
                "kontrakt": str(row[26]),
                "tolov": str(row[27]),
                "qarzi": str(row[28]),
                "ortiqcha": str(row[29]),
            }
            print("Data found at row index:", index)
            print("Data extracted:", data)
            return data

    if not found:
        print(f"'{passport_pin}' not found in column C.")
    return None

password = "1112"


